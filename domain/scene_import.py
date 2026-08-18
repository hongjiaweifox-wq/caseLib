"""
@file scene_import.py
@brief Parse an uploaded Excel(.xlsx)/CSV file into scene control points [{t_min, w}].
@note Two columns expected: 时间(分钟 or HH:MM or Excel time) + 功率(W). Header row auto-detected.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io
from typing import Any, Dict, List, Optional, Tuple

_TIME_KEYS = ("time", "min", "分钟", "时间", "时刻", "hour", "小时", "hh")
_POWER_KEYS = ("power", "watt", "功率", "瓦", "负载", "load")


def _to_minutes(v: Any) -> Optional[float]:
    if isinstance(v, (_dt.time, _dt.datetime)):
        return v.hour * 60 + v.minute + v.second / 60.0
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return f * 1440.0 if 0.0 < f < 1.0 else f  # Excel time = fraction of day
    s = str(v or "").strip()
    if not s:
        return None
    if ":" in s:  # HH:MM[:SS]
        parts = s.split(":")
        try:
            h = int(parts[0])
            m = int(parts[1])
            sec = int(parts[2]) if len(parts) > 2 else 0
            return h * 60 + m + sec / 60.0
        except (ValueError, IndexError):
            return None
    try:
        f = float(s)
        return f * 1440.0 if 0.0 < f < 1.0 else f
    except ValueError:
        return None


def _to_watts(v: Any) -> Optional[float]:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v or "").strip().replace(",", "")
    for suffix in ("W", "w", "瓦"):
        s = s.replace(suffix, "")
    s = s.strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _rows_from_xlsx(raw: bytes) -> List[List[Any]]:
    import openpyxl  # lazy: only needed for xlsx

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _rows_from_csv(raw: bytes) -> List[List[Any]]:
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    delim = "\t" if sample.count("\t") > sample.count(",") else ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=delim)]


def _has_alpha(row: List[Any]) -> bool:
    """A header row carries letters/CJK in a cell that is not an HH:MM time."""
    for c in row:
        if c is None:
            continue
        s = str(c)
        if ":" in s:
            continue
        if any(ch.isalpha() for ch in s):
            return True
    return False


def _pick_columns(rows: List[List[Any]]) -> Tuple[int, int, int]:
    """Return (time_col, power_col, data_start_row)."""
    if not rows:
        return 0, 1, 0
    first = rows[0]
    if not _has_alpha(first):
        return 0, 1, 0  # no header: assume col0=time, col1=power
    tcol, pcol = 0, 1
    for i, c in enumerate(first):
        name = str(c or "").strip().lower()
        if not name:
            continue
        if any(k in name for k in _TIME_KEYS):
            tcol = i
        elif any(k in name for k in _POWER_KEYS):
            pcol = i
    if pcol == tcol:
        pcol = tcol + 1
    return tcol, pcol, 1


def build_import_template() -> bytes:
    """A ready-to-fill .xlsx: col A 时间(HH:MM), col B 功率(W), sample sunny-day curve."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "曲线"
    ws.append(["时间", "功率(W)"])
    # HH:MM every hour + a sample PV-like shape (edit these two columns to your data)
    sample = [0, 0, 0, 0, 0, 0, 40, 160, 320, 470, 560, 600,
              600, 560, 470, 340, 190, 70, 10, 0, 0, 0, 0, 0, 0]
    for h, w in enumerate(sample):
        ws.append([f"{h}:00", w])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    note = ws.cell(row=1, column=4,
                   value="说明：A列=时间(HH:MM 或 分钟 或 Excel时间)，B列=功率W；行数不限，导入按此形状重采样")
    note.font = note.font.copy(italic=True)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_curve_upload(filename: str, raw: bytes) -> Dict[str, Any]:
    """Parse xlsx/csv bytes → {points: [{t_min, w}], ...meta}."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xltx")):
        rows = _rows_from_xlsx(raw)
    else:
        rows = _rows_from_csv(raw)

    tcol, pcol, start = _pick_columns(rows)
    ncols = max((len(r) for r in rows), default=0)
    single_col = ncols < 2

    parsed: List[Tuple[float, float]] = []
    if single_col:
        # one column of power values → spread evenly across 24h
        vals = []
        for r in rows[start if _has_alpha(rows[0]) else 0:]:
            if not r:
                continue
            w = _to_watts(r[0])
            if w is not None:
                vals.append(w)
        n = len(vals)
        for i, w in enumerate(vals):
            t = (i / (n - 1) * 1440.0) if n > 1 else 0.0
            parsed.append((t, w))
    else:
        for r in rows[start:]:
            if not r or len(r) <= max(tcol, pcol):
                continue
            t = _to_minutes(r[tcol])
            w = _to_watts(r[pcol])
            if t is None or w is None:
                continue
            parsed.append((t, w))

    # sort + dedup by rounded minute (last wins)
    dedup: Dict[int, int] = {}
    for t, w in parsed:
        dedup[int(round(t))] = max(0, int(round(w)))
    points = [{"t_min": t, "w": dedup[t]} for t in sorted(dedup)]

    return {
        "points": points,
        "row_count": len(rows),
        "time_col": tcol,
        "power_col": pcol,
        "single_col": single_col,
        "used": len(points),
    }
